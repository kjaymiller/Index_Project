from index_project import missing_items
from openpyxl import Workbook

missing_values = []

for sheet in sheets:
    sheet = wb[sheet]
    for index, row in enumerate(sheet.rows):
        no_item = missing_items((index, row), 2) 
        if no_item:
            missing_values.append((no_item, sheet.title))
            
new_file = "Catalog Missing Items.xlsx"
new_wb = Workbook()
for val in enumerate(missing_values):
    new_wb.active.cell(row=val[0]+1, column=1, value=val[1][0])
    new_wb.active.cell(row=val[0]+1, column=2, value=val[1][1])
new_wb.save(new_file)


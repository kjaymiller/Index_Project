from openpyxl import load_workbook
from index_project import copy_changes_to_sheet
from sys import argv
wb = load_workbook(argv[1], data_only=True)
sheets = wb.sheetnames

src_sheet = wb[sheets[0]]
dest_sheets = [wb[sheet] for sheet in sheets[1:]]

copy_changes_to_sheet(src_sheet, dest_sheets)

wb.save(argv[2])


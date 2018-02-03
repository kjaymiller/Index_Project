from index_project import create_master_list
from openpyxl import load_workbook, Workbook

wb = load_workbook('LKP_Updated_Sheets_Description_per_CS.xlsx', data_only=True)
wb2 = Workbook()
wb2_sheet = wb2.active
sheets = [wb[sheet] for sheet in wb.sheetnames]

master_list = create_master_list(sheets)
for item in enumerate(master_list):
    cell1 = wb2_sheet.cell(row=item[0]+1, column=1)
    cell2 = wb2_sheet.cell(row=item[0]+1, column=2)

    cell1.value = item[1][0]
    cell2.value = item[1][1]

for row in enumerate(wb2_sheet.rows):
    pages = []
    for sheet in sheets:
        for sheet_row in sheet.rows:
            if sheet_row[0].value == row[1][0].value and sheet_row[2].value:
                for page in str(sheet_row[2].value).split(', '):
                    pages.append(int(page))
    pages = sorted([page for page in set(pages)])
    cell = wb2_sheet.cell(row=row[0]+1, column=3)
    cell.value = ', '.join([str(page) for page in pages])
    print(cell.value)
wb2.save('Bulk Master Index.xlsx')

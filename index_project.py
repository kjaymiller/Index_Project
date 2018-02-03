""" FOR EACH FUNCTION YOU WILL NEED TO ADD THE FOLLOWING
for sheet in sheets:
    sheet = wb[sheet]
    for index, row in enumerate(sheet.rows):
"""

from openpyxl import load_workbook, Workbook
import re
from titlecase import titlecase

# Get Rows with Multiple Page Numbers
def concat_rows(enumerated_row, column, row_start, filename):
    """ Takes Rows of Multiple Lengths and Concatenates
    BUG: If first column isn't an int, it will not populate the field.
    """
    
    cell = sheet.cell(row=enumerated_row[0] + 1, column=column)
    index_pages = [str(x.value) for x in enumerated_row[1] if
    any((isinstance(x.value, str), isinstance(x.value, int)))][row_start:]
    return ' '.join(index_pages)
    # wb.save(filename=filename) # Remove and Add to Individual Function


# List all Not-Found Items
def missing_items(sheet, enumerated_row, column):
    cell = sheet.cell(row=enumerated_row[0] + 1, column=column)
    if cell.value == '#N/A':
        cell_value = sheet.cell(row=enumerated_row[0] + 1, column=column - 1).value
        return cell_value


# Convert Field to TitleCase
def titlecase_description(sheet, enumerated_row, column):
   current_cell = sheet.cell(row=enumerated_row[0] + 1, column=column)
   current_value = current_cell.value
   current_cell.value = titlecase(current_value)
   
# Copy Entries from Sheet to Another
def copy_changes_to_sheet(src_sheet, dest_sheets):
    for row in src_sheet.rows:
        item_no = row[0].value
        for dest_sheet in dest_sheets:
            for dest_sheet_row in dest_sheet.rows:
                if item_no == dest_sheet_row[0].value:
                    print(f'{dest_sheet_row[0].value} - {row[1].value}')
                    dest_sheet_row[1].value = row[1].value
                    break


def create_master_list(src_sheets):
    item_list_master = []

    for sheet in src_sheets:
        for row in sheet:
            item_list_master.append((row[0].value, row[1].value))

    return set(item_list_master)
